"""
Excel 格式化器
"""

from typing import List, Dict, Any
from io import BytesIO

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from .base import BaseFormatter


class ExcelFormatter(BaseFormatter):
    """
    Excel 格式化器
    
    功能：
    1. 写入 Excel (.xlsx) 文件
    2. 支持样式设置
    3. 分批写入
    """
    
    def __init__(
        self,
        fields: List[str],
        field_names: List[str] = None,
        header_style: bool = True
    ):
        """
        初始化 Excel 格式化器
        
        Args:
            fields: 导出字段列表
            field_names: 字段显示名称（可选）
            header_style: 是否使用表头样式
        """
        if not EXCEL_AVAILABLE:
            raise ImportError(
                "openpyxl library is required for Excel export. "
                "Install it with: pip install openpyxl"
            )
        
        super().__init__(fields)
        self.field_names = field_names or fields
        self.header_style = header_style
        self.workbook = None
        self.worksheet = None
        self.current_row = 1
    
    def open(self, file_path: str):
        """
        创建 Excel 文件
        
        Args:
            file_path: 文件路径
        """
        self.file_path = file_path
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Export Data"
        
        # 写入表头
        for col, field_name in enumerate(self.field_names, 1):
            cell = self.worksheet.cell(row=1, column=col, value=field_name)
            
            if self.header_style:
                # 设置表头样式
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="4472C4",
                    end_color="4472C4",
                    fill_type="solid"
                )
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        self.current_row = 2
        self.row_count = 0
    
    def write_batch(self, data: List[Dict[str, Any]]):
        """
        写入一批数据
        
        Args:
            data: 数据列表
        """
        if not self.worksheet:
            raise RuntimeError("File not opened. Call open() first.")
        
        for row_data in data:
            values = self._extract_values(row_data)
            
            for col, value in enumerate(values, 1):
                cell = self.worksheet.cell(
                    row=self.current_row,
                    column=col,
                    value=value
                )
                
                # 自动调整日期和数字格式
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal='right')
            
            self.current_row += 1
            self.row_count += 1
    
    def close(self):
        """关闭并保存文件"""
        if self.workbook and self.file_path:
            # 自动调整列宽
            for col in range(1, len(self.fields) + 1):
                self.worksheet.column_dimensions[
                    self._get_column_letter(col)
                ].auto_size = True
            
            # 保存文件
            self.workbook.save(self.file_path)
            self.workbook = None
            self.worksheet = None
    
    def _get_column_letter(self, col_num: int) -> str:
        """
        将列号转换为 Excel 列字母
        
        Args:
            col_num: 列号（1-based）
        
        Returns:
            列字母（A, B, ..., Z, AA, AB, ...）
        """
        from openpyxl.utils import get_column_letter
        return get_column_letter(col_num)
    
    def get_stats(self) -> dict:
        """
        获取统计信息
        
        Returns:
            统计数据
        """
        return {
            'file_path': self.file_path,
            'row_count': self.row_count,
            'field_count': len(self.fields)
        }
